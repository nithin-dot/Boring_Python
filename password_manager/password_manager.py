import random
import openpyxl
from openpyxl.reader.excel import load_workbook
import string
import xlrd
import getpass 

def get_pass():
    filled=0
    word=0
    columns=0
    book =openpyxl.load_workbook('D:/python/automation_scripts/password_manager/schedule.xlsx') 
    sheet1 = book['Sheet1']     
    s="abcdefghijklmnopqrstuvwxyz0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ!@#$%^&*()?/.,><:;{}[]'\'"
    wb=xlrd.open_workbook("schedule.xlsx")
    sheet=wb.sheet_by_index(0)
    for row in range (sheet.nrows):
        for column in range (sheet.ncols) :
            if (sheet.cell_value(row,column)==""):
                filled
            else :
                filled+=1
    filled=int(filled/3)+1
    Domain=input("Enter the domain Name:")
    for row in range (sheet.nrows):
        columns+=1
        for column in range (sheet.ncols) :
            if (sheet.cell_value(row,column)==Domain):
             word=columns
    if(word<2):
        Email=input('''Enter the Email Id
        1) nithinkirthick@hotmail.com
        2) nithinkirthick@gmail.com
        3) nithinkirthick.it19@bitsathy.ac.in
        Enter the id : ''') 
        len_pwd=int(input("Enter number of letters in password:"))
        s="".join(random.sample(s,len_pwd))
        encstr=""
        for i in s:
            if(ord(i))>=33 and (ord(i)<=64):
                temp=(ord(i)+26)
                if temp>64:
                    temp=temp%64+32
                encstr=encstr+chr(temp)
            elif(ord(i))>=65 and (ord(i)<=96):
                temp=(ord(i)+17)
                if temp>96:
                    temp=temp%96+64        
                encstr=encstr+chr(temp)
            elif(ord(i))>=97 and (ord(i)<=122):
                temp=(ord(i)+11)
                if temp>122:
                    temp=temp%122+96
                encstr=encstr+chr(temp)
            elif(ord(i))>=123 and (ord(i)<=126):
                temp=(ord(i)+3)
                if temp>126:
                    temp=temp%126+122
                encstr=encstr+chr(temp)
            else:
                encstr=encstr+chr(ord(i))
        print(s)
        val1=sheet1.cell(filled,1)
        val2=sheet1.cell(filled,2)
        val3=sheet1.cell(filled,3)
        val1.value=Domain
        val2.value=Email
        val3.value=encstr
        book.save('schedule.xlsx')
        input()
    else:
        vali1=sheet1['B'+str(word)].value
        vali2=sheet1['C'+str(word)].value
        decstr=""
        for i in vali2:
            if(ord(i))>=33 and (ord(i)<=64):
                decstr=decstr+chr((ord(i)-26-33)%32+33)
            elif((ord(i))>=65) and (ord(i))<=96:
                decstr=decstr+chr((ord(i) -17-65) % 32 + 65)
            elif((ord(i))>=97) and (ord(i))<=122:
                decstr=decstr+chr((ord(i) - 11 - 97) % 26 + 97)
            elif(ord(i))>=123 and (ord(i)<=126):
                 decstr=decstr+chr((ord(i)-3-123)%4+123)
            else:
                decstr= decstr+chr(ord(i))
        print(vali1)
        print(decstr)
        input()
if __name__ == "__main__":
   p = getpass.getpass() 
   if("1" in p)and(p.isalnum()==False)and(p.endswith('N', 0, 1)):
     my_inbox = get_pass()
