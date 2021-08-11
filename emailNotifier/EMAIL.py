import imaplib
import openpyxl
import gtts
from playsound import playsound
import re
import xlsxwriter    
from pytz import timezone
import dateutil.parser
from email.utils import parsedate_tz, mktime_tz, formatdate
from datetime import timedelta
import email
import time
import datetime
book =openpyxl.load_workbook('schedule.xlsx') 
sheet = book['Sheet1']     
host = 'imap.gmail.com'
username = 'example@example'
password = 'password'
date_format = datetime.datetime.now() 
byte=[]
def get_inbox():
    mail = imaplib.IMAP4_SSL(host)
    mail.login(username, password)
    status, messages = mail.select("INBOX",readonly=True)
    today = datetime.datetime.now()
    yesterday = today - timedelta(days = 1)  
    Time=yesterday.strftime("%d-%b-%Y")
    date='SINCE '+Time
    _,search_data=mail.search(None,date)
    # data = mail.fetch('3968','RFC822')
    # print(data)
    messages = int(messages[0])
    print(messages)
    my_message=[]
    row = 2    
    column = 1  
    columns = 2
    for num in reversed(search_data[0].split()):
        print(num)
    #   if date_format.strftime("%a, %b %d")==suc:
        email_data = {}
        _, data = mail.fetch(num,'(RFC822)')
        # print(data[0])
        _, b = data[0]
        email_message = email.message_from_bytes(b)
        for header in ['from']:
            val=sheet.cell(row,column)
            val.value=email_message[header].rsplit('<', 1)[0]
            # tts = gtts.gTTS("mail from"+email_message[header].rsplit('<', 1)[0])
            # tts.save("C:/Users/nithin/Desktop/email/"+num.decode()+".mp3")
            # playsound("C:/Users/nithin/Desktop/email/"+num.decode()+".mp3")    
            # st = dateutil.parser.parse(email_message['date'])
            print("{}: {}".format(header, email_message[header]))
            # print(st.strftime('%a, %b %d, %Y at %I:%M %p'))
            email_data[header] = email_message[header]
        for part in email_message.walk():
            if part.get_content_type() == "text/plain":
                body = part.get_payload(decode=True)
                email_data['body'] = body.decode()
                regax='http[s]?://forms.gle/(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
                # regax='http[s]?://meet.google.com/(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
                match=re.findall(regax, email_data['body'])
                # match = re.findall(regaxs, email_data['body'])
                # re.findall(regex, email_data['body'])
                # match = re.findall(r"on \d.*\d", email_data['body'])
                # word = ['AM','PM','Pm','pM','pm','2020']
                # text=email_data['body']
                # print(text[text.index('on')+len('on'):text.index('.')])
                for m in match:                 
                    data=sheet.cell(row,columns)
                    data.value=m
                    row+=1   
            elif part.get_content_type() == "text/html":
                html_body = part.get_payload(decode=True)
                email_data['html_body'] = html_body.decode()
        my_message.append(email_data)
    return my_message
    


if __name__ == "__main__":
    my_inbox = get_inbox()
    book.save('schedule.xlsx')
#    print(search_data)